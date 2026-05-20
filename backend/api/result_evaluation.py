"""
Роутер для работы с результатами оценки
"""
from typing import List, Optional
from decimal import Decimal
import asyncio
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, status, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, desc
from sqlalchemy.orm import selectinload
from pydantic import BaseModel, Field

from backend.db.database import get_db
from backend.models.models import (
    ResultEvaluation,
    ResultCategoryScore,
    Resume,
    Model,
    UserAccount,
    RecommendationRule,
    EvaluationCategory,
    Candidate,
    ResumeVacancy,
    Vacancy,
)
from backend.schemas.result_evaluation import (
    ResultEvaluationCreate,
    ResultEvaluationUpdate,
    ResultEvaluationResponse,
    ResultEvaluationDetailed,
    ResultEvaluationListItem,
    CandidateInfo,
    VacancyInfo,
    CategoryScoreDetailed,
)

router = APIRouter()


# =============================================================================
# Dependencies
# =============================================================================

def get_evaluator(request: Request):
    """
    Dependency для получения ML evaluator из app state

    Модель загружается при старте сервера (см. lifespan в main.py)
    """
    evaluator = request.app.state.evaluator

    if evaluator is None:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail=(
                "ML модель не загружена. "
                "Проверьте логи сервера при запуске - возможно модель не найдена "
                "или не установлены библиотеки (torch, sentence-transformers)."
            )
        )

    return evaluator


# =============================================================================
# Схемы для запуска оценки
# =============================================================================

class EvaluateRequest(BaseModel):
    """Запрос на запуск оценки"""
    vacancy_id: int = Field(..., gt=0, description="ID вакансии")
    resume_ids: Optional[List[int]] = Field(None, description="ID резюме (опционально, если не указано - все резюме по вакансии)")


class EvaluateResultItem(BaseModel):
    """Элемент результата оценки"""
    result_id: int
    resume_id: int
    candidate_name: str
    overall_score: Decimal
    recommendation: str


class EvaluateResponse(BaseModel):
    """Ответ на запрос оценки"""
    message: str
    count: int
    results: List[EvaluateResultItem]


async def determine_recommendation(overall_score: Decimal, db: AsyncSession) -> int:
    """Определить recommendation_id на основе overall_score"""
    result = await db.execute(select(RecommendationRule))
    rules = result.scalars().all()

    for rule in rules:
        if rule.min_score <= overall_score <= rule.max_score:
            return rule.recommendation_id

    # По умолчанию - самая низкая рекомендация
    lowest_rule = min(rules, key=lambda r: r.min_score)
    return lowest_rule.recommendation_id


@router.post("/evaluate", response_model=EvaluateResponse, status_code=status.HTTP_200_OK)
async def evaluate_resumes(
    request: EvaluateRequest,
    db: AsyncSession = Depends(get_db),
    evaluator = Depends(get_evaluator)
):
    """
    Запустить оценку резюме для вакансии

    1. Получает все резюме по vacancy_id (или указанные в resume_ids)
    2. Загружает текст резюме и описание вакансии
    3. Вызывает ResumeEvaluator для оценки
    4. Сохраняет результаты в result_evaluation и result_category_score
    5. Возвращает список результатов
    """
    # Проверка существования вакансии
    vacancy = await db.get(Vacancy, request.vacancy_id)
    if not vacancy:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Вакансия с ID {request.vacancy_id} не найдена"
        )

    if not vacancy.description:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="У вакансии отсутствует описание"
        )

    # Получение резюме
    if request.resume_ids:
        # Указаны конкретные резюме
        resumes_query = (
            select(Resume)
            .options(selectinload(Resume.candidate))
            .join(ResumeVacancy)
            .where(
                ResumeVacancy.vacancy_id == request.vacancy_id,
                Resume.resume_id.in_(request.resume_ids)
            )
        )
    else:
        # Все резюме по вакансии
        resumes_query = (
            select(Resume)
            .options(selectinload(Resume.candidate))
            .join(ResumeVacancy)
            .where(ResumeVacancy.vacancy_id == request.vacancy_id)
        )

    result_query = await db.execute(resumes_query)
    resumes = result_query.scalars().all()

    if not resumes:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Резюме для вакансии {request.vacancy_id} не найдены"
        )

    # Получение модели (берём первую активную модель или последнюю)
    model_result = await db.execute(
        select(Model).order_by(Model.finetune_date.desc()).limit(1)
    )
    model = model_result.scalar_one_or_none()

    if not model:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="ML-модель не найдена в БД. Необходимо сначала обучить модель."
        )

    # TODO: Получить текущего пользователя из JWT токена
    # Пока используем первого пользователя
    user_result = await db.execute(select(UserAccount).limit(1))
    user = user_result.scalar_one_or_none()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Пользователь не найден"
        )

    # Получение категорий оценки
    categories_result = await db.execute(select(EvaluationCategory))
    categories = categories_result.scalars().all()
    category_map = {cat.name: cat.category_id for cat in categories}

    if not categories:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Категории оценки не найдены в БД"
        )

    print("1. Загрузка категорий завершена")
    print("2. Модель готова к использованию (загружена при старте приложения)")

    # Оценка каждого резюме
    results = []
    vacancy_text = vacancy.description
    loop = asyncio.get_event_loop()

    # Создаём ThreadPoolExecutor один раз для всех ML вычислений
    with ThreadPoolExecutor() as executor:
        for resume in resumes:
            # Проверка наличия текста резюме
            if not resume.extracted_text:
                # Пропускаем резюме без текста
                continue

            print(f"3. Начало оценки резюме {resume.resume_id}")

            # Вызов ResumeEvaluator в отдельном потоке
            # чтобы не блокировать async event loop
            eval_result = await loop.run_in_executor(
                executor,
                lambda rt=resume.extracted_text, vt=vacancy_text: evaluator.evaluate(
                    resume_text=rt,
                    vacancy_text=vt,
                    return_categories=True
                )
            )

            print(f"4. Оценка завершена {resume.resume_id}")

        overall_score = Decimal(str(eval_result["overall_score"]))
        category_scores = eval_result.get("category_scores", {})

        # Определение рекомендации
        recommendation_id = await determine_recommendation(overall_score, db)

        # Создание записи результата оценки
        result_evaluation = ResultEvaluation(
            overall_score=overall_score,
            resume_id=resume.resume_id,
            model_id=model.model_id,
            user_id=user.user_id,
            recommendation_id=recommendation_id,
        )
        db.add(result_evaluation)
        await db.flush()  # Получаем result_id

        # Сохранение оценок по категориям
        for category_name, score in category_scores.items():
            if category_name in category_map:
                category_score = ResultCategoryScore(
                    result_id=result_evaluation.result_id,
                    category_id=category_map[category_name],
                    score=Decimal(str(score)),
                )
                db.add(category_score)

        await db.flush()

        print(f"5. Сохранение в БД завершено для резюме {resume.resume_id}")

        # Получение recommendation_text
        recommendation = await db.get(RecommendationRule, recommendation_id)

        # Формирование имени кандидата
        candidate = resume.candidate
        candidate_name = f"{candidate.last_name} {candidate.first_name}"
        if candidate.patronymic:
            candidate_name += f" {candidate.patronymic}"

        results.append(
            EvaluateResultItem(
                result_id=result_evaluation.result_id,
                resume_id=resume.resume_id,
                candidate_name=candidate_name,
                overall_score=overall_score,
                recommendation=recommendation.recommendation_text if recommendation else ""
            )
        )

    await db.commit()

    print("6. Всё завершено")

    # Сортировка результатов по убыванию балла
    results.sort(key=lambda x: x.overall_score, reverse=True)

    return EvaluateResponse(
        message=f"Оценка завершена для {len(results)} резюме",
        count=len(results),
        results=results
    )


@router.post("/", response_model=ResultEvaluationResponse, status_code=status.HTTP_201_CREATED)
async def create_evaluation(
    evaluation_data: ResultEvaluationCreate,
    db: AsyncSession = Depends(get_db)
):
    """Создать результат оценки резюме"""
    # Проверка существования связанных сущностей
    resume = await db.get(Resume, evaluation_data.resume_id)
    if not resume:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Резюме с ID {evaluation_data.resume_id} не найдено"
        )

    model = await db.get(Model, evaluation_data.model_id)
    if not model:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Модель с ID {evaluation_data.model_id} не найдена"
        )

    user = await db.get(UserAccount, evaluation_data.user_id)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Пользователь с ID {evaluation_data.user_id} не найден"
        )

    # Определение рекомендации на основе overall_score
    recommendation_id = await determine_recommendation(evaluation_data.overall_score, db)

    # Создание результата оценки
    result = ResultEvaluation(
        overall_score=evaluation_data.overall_score,
        resume_id=evaluation_data.resume_id,
        model_id=evaluation_data.model_id,
        user_id=evaluation_data.user_id,
        recommendation_id=recommendation_id,
    )
    db.add(result)
    await db.flush()  # Получаем result_id до commit

    # Создание оценок по категориям
    for category_score in evaluation_data.category_scores:
        # Проверка существования категории
        category = await db.get(EvaluationCategory, category_score.category_id)
        if not category:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Категория с ID {category_score.category_id} не найдена"
            )

        score = ResultCategoryScore(
            result_id=result.result_id,
            category_id=category_score.category_id,
            score=category_score.score,
        )
        db.add(score)

    await db.commit()
    await db.refresh(result)

    return result


@router.get("/", response_model=List[ResultEvaluationListItem])
async def get_evaluations(
    skip: int = 0,
    limit: int = 100,
    resume_id: int = None,
    vacancy_id: int = None,
    db: AsyncSession = Depends(get_db)
):
    """Получить список всех результатов оценки (с фильтрацией по resume_id или vacancy_id)"""
    query = (
        select(ResultEvaluation)
        .options(
            selectinload(ResultEvaluation.resume).selectinload(Resume.candidate),
            selectinload(ResultEvaluation.resume).selectinload(Resume.vacancy_applications),
            selectinload(ResultEvaluation.recommendation)
        )
        .join(Resume)
        .join(Candidate)
        .join(RecommendationRule)
        .outerjoin(ResumeVacancy)
        .outerjoin(Vacancy)
    )

    if resume_id is not None:
        query = query.where(ResultEvaluation.resume_id == resume_id)

    if vacancy_id is not None:
        query = query.where(ResumeVacancy.vacancy_id == vacancy_id)

    query = query.order_by(desc(ResultEvaluation.overall_score)).offset(skip).limit(limit)
    result_query = await db.execute(query)
    results = result_query.scalars().all()

    # Формирование списка
    evaluations = []
    for result in results:
        candidate = result.resume.candidate
        candidate_name = f"{candidate.last_name} {candidate.first_name}"
        if candidate.patronymic:
            candidate_name += f" {candidate.patronymic}"

        evaluations.append(
            ResultEvaluationListItem(
                result_id=result.result_id,
                overall_score=result.overall_score,
                analysis_date=result.analysis_date,
                candidate_name=candidate_name,
                vacancy_title=None,  # Можно получить отдельным запросом если нужно
                recommendation_text=result.recommendation.recommendation_text,
            )
        )

    return evaluations


@router.get("/{result_id}", response_model=ResultEvaluationDetailed)
async def get_evaluation(
    result_id: int,
    db: AsyncSession = Depends(get_db),
    evaluator = Depends(get_evaluator)
):
    """Получить детальный результат оценки по ID"""
    # Загрузка результата с eager loading всех relationships
    stmt = (
        select(ResultEvaluation)
        .options(
            selectinload(ResultEvaluation.resume).selectinload(Resume.candidate),
            selectinload(ResultEvaluation.resume).selectinload(Resume.vacancy_applications).selectinload(ResumeVacancy.vacancy),
            selectinload(ResultEvaluation.recommendation),
            selectinload(ResultEvaluation.user),
            selectinload(ResultEvaluation.model),
            selectinload(ResultEvaluation.category_scores).selectinload(ResultCategoryScore.category)
        )
        .where(ResultEvaluation.result_id == result_id)
    )
    result_query = await db.execute(stmt)
    result = result_query.scalar_one_or_none()

    if not result:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Результат оценки с ID {result_id} не найден"
        )

    # Получение кандидата
    candidate = result.resume.candidate
    candidate_full_name = f"{candidate.last_name} {candidate.first_name}"
    if candidate.patronymic:
        candidate_full_name += f" {candidate.patronymic}"

    # Получение вакансии (берём первую из vacancy_applications)
    vacancy = None
    if result.resume.vacancy_applications:
        vacancy = result.resume.vacancy_applications[0].vacancy

    if not vacancy:
        # Если нет вакансии в resume_vacancy, создаём пустую
        vacancy_info = {
            "vacancy_id": 0,
            "title": "Не указана",
            "description": None,
            "specialization_id": None,
            "is_active": False
        }
    else:
        vacancy_info = {
            "vacancy_id": vacancy.vacancy_id,
            "title": vacancy.title,
            "description": vacancy.description,
            "specialization_id": vacancy.specialization_id,
            "is_active": vacancy.is_active
        }

    # Формирование оценок по категориям
    category_scores_detailed = []
    for cs in result.category_scores:
        weight = cs.category.weight if cs.category.weight else Decimal("0.0")
        weighted_score = (cs.score * weight) / Decimal("100.0")

        category_scores_detailed.append(
            CategoryScoreDetailed(
                category_name=cs.category.name,
                score=cs.score,
                weight=weight,
                weighted_score=weighted_score
            )
        )

    # Формирование анализа соответствия
    matching_analysis = None
    if (vacancy and vacancy.description and
        result.resume.extracted_text):
        # Запуск анализа в отдельном потоке
        loop = asyncio.get_event_loop()
        with ThreadPoolExecutor() as executor:
            analysis_result = await loop.run_in_executor(
                executor,
                lambda: evaluator.analyze_matching(
                    resume_text=result.resume.extracted_text,
                    vacancy_text=vacancy.description
                )
            )

        from backend.schemas.result_evaluation import MatchingAnalysis
        matching_analysis = MatchingAnalysis(**analysis_result)

    # Формирование детального ответа
    detailed_response = ResultEvaluationDetailed(
        id=result.result_id,
        candidate=CandidateInfo(
            candidate_id=candidate.candidate_id,
            last_name=candidate.last_name,
            first_name=candidate.first_name,
            patronymic=candidate.patronymic,
            email=candidate.email,
            phone=candidate.phone,
            full_name=candidate_full_name
        ),
        vacancy=VacancyInfo(**vacancy_info),
        overall_score=result.overall_score,
        category_scores=category_scores_detailed,
        recommendation=result.recommendation.recommendation_text,
        evaluated_at=result.analysis_date,
        evaluator_name=result.user.login,
        resume_url=result.resume.file_path if result.resume.file_path else None,
        matching_analysis=matching_analysis
    )

    return detailed_response


@router.put("/{result_id}", response_model=ResultEvaluationResponse)
async def update_evaluation(
    result_id: int,
    evaluation_data: ResultEvaluationUpdate,
    db: AsyncSession = Depends(get_db)
):
    """Обновить результат оценки"""
    result = await db.get(ResultEvaluation, result_id)
    if not result:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Результат оценки с ID {result_id} не найден"
        )

    # Если обновляется overall_score, пересчитываем recommendation_id
    update_data = evaluation_data.model_dump(exclude_unset=True)

    if "overall_score" in update_data:
        recommendation_id = await determine_recommendation(update_data["overall_score"], db)
        update_data["recommendation_id"] = recommendation_id

    # Обновление полей
    for field, value in update_data.items():
        setattr(result, field, value)

    await db.commit()
    await db.refresh(result)

    return result


@router.delete("/{result_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_evaluation(
    result_id: int,
    db: AsyncSession = Depends(get_db)
):
    """Удалить результат оценки"""
    result = await db.get(ResultEvaluation, result_id)
    if not result:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Результат оценки с ID {result_id} не найден"
        )

    # Удаление связанных оценок по категориям (CASCADE)
    await db.execute(
        select(ResultCategoryScore).where(ResultCategoryScore.result_id == result_id)
    )

    db.delete(result)
    await db.commit()

    return None


class UpdateRecommendationRequest(BaseModel):
    """Запрос на обновление рекомендации"""
    recommendation_text: str = Field(..., description="Текст рекомендации")


@router.patch("/{result_id}/recommendation", response_model=ResultEvaluationResponse)
async def update_recommendation(
    result_id: int,
    request: UpdateRecommendationRequest,
    db: AsyncSession = Depends(get_db)
):
    """
    Обновить рекомендацию для результата оценки вручную
    """
    # Получить результат
    result = await db.get(ResultEvaluation, result_id)
    if not result:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Результат оценки с ID {result_id} не найден"
        )

    # Найти рекомендацию по тексту
    recommendation_query = await db.execute(
        select(RecommendationRule).where(
            RecommendationRule.recommendation_text == request.recommendation_text
        )
    )
    recommendation = recommendation_query.scalar_one_or_none()

    if not recommendation:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Рекомендация '{request.recommendation_text}' не найдена"
        )

    # Обновить рекомендацию
    result.recommendation_id = recommendation.recommendation_id
    await db.commit()
    await db.refresh(result)

    return result


@router.get("/export/{vacancy_id}")
async def export_results(
    vacancy_id: int,
    format: str = "excel",  # pdf или excel
    db: AsyncSession = Depends(get_db)
):
    """
    Экспорт результатов оценки для вакансии в PDF или Excel
    """
    # Получение всех результатов для вакансии
    query = (
        select(ResultEvaluation)
        .options(
            selectinload(ResultEvaluation.resume).selectinload(Resume.candidate),
            selectinload(ResultEvaluation.recommendation),
            selectinload(ResultEvaluation.category_scores).selectinload(ResultCategoryScore.category)
        )
        .join(Resume)
        .join(ResumeVacancy)
        .where(ResumeVacancy.vacancy_id == vacancy_id)
        .order_by(desc(ResultEvaluation.overall_score))
    )

    result_query = await db.execute(query)
    results = result_query.scalars().all()

    if not results:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Результаты для вакансии {vacancy_id} не найдены"
        )

    # Получение информации о вакансии
    vacancy = await db.get(Vacancy, vacancy_id)
    vacancy_title = vacancy.title if vacancy else f"Вакансия #{vacancy_id}"

    if format == "excel":
        return await export_to_excel(results, vacancy_title)
    elif format == "pdf":
        return await export_to_pdf(results, vacancy_title)
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Неподдерживаемый формат. Используйте 'pdf' или 'excel'"
        )


async def export_to_excel(results: List[ResultEvaluation], vacancy_title: str):
    """Экспорт результатов в Excel (XLSX)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Если openpyxl не установлен, возвращаем CSV
        return await export_to_csv(results, vacancy_title)

    # Создание workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты оценки"

    # Заголовок
    ws['A1'] = f"Результаты оценки для вакансии: {vacancy_title}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')

    ws['A2'] = f"Дата экспорта: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws.merge_cells('A2:F2')

    # Заголовки таблицы
    headers = ["№", "ФИО кандидата", "Общий балл", "Рекомендация", "Дата оценки", "Детали по категориям"]
    ws.append([])  # Пустая строка
    ws.append(headers)

    # Стилизация заголовков
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Заполнение данных
    for idx, result in enumerate(results, 1):
        candidate = result.resume.candidate
        candidate_name = f"{candidate.last_name} {candidate.first_name}"
        if candidate.patronymic:
            candidate_name += f" {candidate.patronymic}"

        # Формирование строки с категориями
        categories_str = "; ".join([
            f"{cs.category.name}: {float(cs.score):.2f}"
            for cs in result.category_scores
        ])

        row = [
            idx,
            candidate_name,
            float(result.overall_score),
            result.recommendation.recommendation_text,
            result.analysis_date.strftime('%d.%m.%Y %H:%M'),
            categories_str
        ]
        ws.append(row)

        # Цветовое кодирование по баллу
        score_cell = ws.cell(row=4+idx, column=3)
        score = float(result.overall_score)
        if score >= 75:
            score_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif score >= 50:
            score_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            score_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Автоматическая ширина столбцов
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)

        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Сохранение в память
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Возврат файла
    headers = {
        'Content-Disposition': f'attachment; filename="results_{vacancy_title[:30]}.xlsx"'
    }
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )


async def export_to_csv(results: List[ResultEvaluation], vacancy_title: str):
    """Экспорт результатов в CSV (для случая когда нет openpyxl)"""
    import csv

    output = BytesIO()
    # Используем utf-8-sig для корректного отображения кириллицы в Excel
    output.write('﻿'.encode('utf-8'))  # BOM для UTF-8

    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)

    # Заголовок
    writer.writerow([f"Результаты оценки для вакансии: {vacancy_title}"])
    writer.writerow([f"Дата экспорта: {datetime.now().strftime('%d.%m.%Y %H:%M')}"])
    writer.writerow([])

    # Заголовки таблицы
    writer.writerow(["№", "ФИО кандидата", "Общий балл", "Рекомендация", "Дата оценки", "Детали по категориям"])

    # Данные
    for idx, result in enumerate(results, 1):
        candidate = result.resume.candidate
        candidate_name = f"{candidate.last_name} {candidate.first_name}"
        if candidate.patronymic:
            candidate_name += f" {candidate.patronymic}"

        categories_str = "; ".join([
            f"{cs.category.name}: {float(cs.score):.2f}"
            for cs in result.category_scores
        ])

        writer.writerow([
            idx,
            candidate_name,
            f"{float(result.overall_score):.2f}",
            result.recommendation.recommendation_text,
            result.analysis_date.strftime('%d.%m.%Y %H:%M'),
            categories_str
        ])

    output.seek(0)

    headers = {
        'Content-Disposition': f'attachment; filename="results_{vacancy_title[:30]}.csv"'
    }
    return StreamingResponse(
        output,
        media_type="text/csv; charset=utf-8",
        headers=headers
    )


async def export_to_pdf(results: List[ResultEvaluation], vacancy_title: str):
    """Экспорт результатов в PDF"""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="Библиотека reportlab не установлена. Используйте формат 'excel'"
        )

    # Создание PDF в памяти
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4))

    # Стили
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=12,
        alignment=1  # Центрирование
    )

    # Элементы документа
    elements = []

    # Заголовок
    title = Paragraph(f"Результаты оценки для вакансии:<br/>{vacancy_title}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 12))

    # Дата экспорта
    date_text = Paragraph(
        f"Дата экспорта: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        styles['Normal']
    )
    elements.append(date_text)
    elements.append(Spacer(1, 20))

    # Таблица с результатами
    table_data = [["№", "ФИО кандидата", "Балл", "Рекомендация", "Дата оценки"]]

    for idx, result in enumerate(results, 1):
        candidate = result.resume.candidate
        candidate_name = f"{candidate.last_name} {candidate.first_name}"
        if candidate.patronymic:
            candidate_name += f" {candidate.patronymic}"

        table_data.append([
            str(idx),
            candidate_name,
            f"{float(result.overall_score):.2f}",
            result.recommendation.recommendation_text,
            result.analysis_date.strftime('%d.%m.%Y')
        ])

    # Создание таблицы
    table = Table(table_data, colWidths=[30, 200, 60, 150, 80])

    # Стилизация таблицы
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    # Цветовое кодирование по баллу
    for idx, result in enumerate(results, 1):
        score = float(result.overall_score)
        if score >= 75:
            bg_color = colors.HexColor('#C6EFCE')
        elif score >= 50:
            bg_color = colors.HexColor('#FFEB9C')
        else:
            bg_color = colors.HexColor('#FFC7CE')

        table.setStyle(TableStyle([
            ('BACKGROUND', (2, idx), (2, idx), bg_color)
        ]))

    elements.append(table)

    # Генерация PDF
    try:
        doc.build(elements)
        output.seek(0)

        headers = {
            'Content-Disposition': f'attachment; filename="results_{vacancy_title[:30]}.pdf"'
        }
        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers=headers
        )
    except Exception as e:
        # Если ошибка при генерации PDF (например, с кириллицей), возвращаем Excel
        print(f"Ошибка генерации PDF: {e}")
        return await export_to_excel(results, vacancy_title)
